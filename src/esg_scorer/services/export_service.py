from typing import List
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..models.schemas import CompanyESGResult

class ExcelExporter:
    @staticmethod
    def export_batch_results(results: List[CompanyESGResult], output_path: str):
        """Xuất danh sách kết quả ESG ra file Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "ESG Results Summary"
        
        # Headers
        headers = [
            "Công ty", "Năm", "Total ESG Score", "Trọng số E", "Trọng số S", "Trọng số G",
             "E Score", "S Score", "G Score",
             "ENV Net", "COM Net", "EMP Net", "DIV Net", "CGOV Net", "PRO Net",
             "ENV Str", "ENV Con", "COM Str", "COM Con", "EMP Str", "EMP Con",
             "DIV Str", "DIV Con", "CGOV Str", "CGOV Con", "PRO Str", "PRO Con"
        ]
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Dữ liệu
        for row, r in enumerate(results, 2):
            comps = r.components
            ws.append([
                r.company_name, r.year, round(r.total_esg_score, 2),
                r.weights.e_weight, r.weights.s_weight, r.weights.g_weight,
                r.e_score, r.s_score, r.g_score,
                # Net scores
                comps["Environment"].net_score, comps["Community"].net_score, 
                comps["Employee Relations"].net_score, comps["Diversity"].net_score,
                comps["Corporate Governance"].net_score, comps["Product Quality"].net_score,
                # Breakdown
                comps["Environment"].strengths.total_score, comps["Environment"].concerns.total_score,
                comps["Community"].strengths.total_score, comps["Community"].concerns.total_score,
                comps["Employee Relations"].strengths.total_score, comps["Employee Relations"].concerns.total_score,
                comps["Diversity"].strengths.total_score, comps["Diversity"].concerns.total_score,
                comps["Corporate Governance"].strengths.total_score, comps["Corporate Governance"].concerns.total_score,
                comps["Product Quality"].strengths.total_score, comps["Product Quality"].concerns.total_score
            ])
            
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        # Tạo sheet chi tiết bằng chứng (Evidence)
        ws_details = wb.create_sheet(title="Evidences")
        ws_details.append(["Công ty", "Thành phần", "Tiêu chí", "ID", "Điểm", "Bằng chứng trích xuất"])
        for col in range(1, 7):
            ws_details.cell(row=1, column=col).font = header_font
            ws_details.cell(row=1, column=col).fill = header_fill
            
        detail_row = 2
        for r in results:
            for comp in r.components.values():
                for item in comp.strengths.items + comp.concerns.items:
                    if item.score > 0:
                        ws_details.append([
                            r.company_name, comp.name.value, item.name, item.id, item.score, item.evidence or ""
                        ])
                        detail_row += 1
                        
        ws_details.column_dimensions["A"].width = 20
        ws_details.column_dimensions["F"].width = 80
        
        # Lưu file
        os_path = Path(output_path)
        os_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
